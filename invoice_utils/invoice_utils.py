import glob
import os
from datetime import datetime
from time import sleep

from openpyxl import load_workbook


def find_text(temp_df, text, case=False):
    location = temp_df[temp_df['TEXT'].str.contains(text, case=case, na=False)].index.tolist()
    return location


def find_next_word(temp_df, text, nearest_text, case=False):
    pos = find_text(temp_df, text, case=case)
    pos = [item + 1 for item in pos]
    #     print(pos)
    new_pos = list()
    for item in pos:
        #         temp = df.loc[]
        temp_value = temp_df.loc[item]['TEXT']
        #         print(temp_value.lower())

        if temp_value.lower().__contains__(nearest_text.lower()):
            new_pos.append(item - 1)

    return new_pos


def get_max_file(folder_path):
    file_type = '\*.xlsx'
    files = glob.glob(folder_path + file_type)
    max_file = max(files, key=os.path.getctime)
    return max_file

def write_excel_sheet(prev_filename, result, supplier_name, sr_no=1):


    wb = load_workbook(filename=f'{prev_filename}')
    dirname = os.path.dirname(prev_filename)
    time_stamp = datetime.now().strftime('%d_%m_%Y-%H_%M_%S')
    sheet = wb.active
    sheet_row = sheet.max_row + 1

    # try:
    item_col = 5
    try:
        if len(result['items'])==0:
            # print('Result _items :', len(result['items']))
            with open("error_pdf/error_log.txt", 'a') as fp:
                fp.write(f"Error in pdf : {supplier_name}")
                fp.write("\n")
    except Exception as e:
        print("""
                1. Error in Write to Error Log file
                2. May be File not Found 
                """)


    for i, item in enumerate(result['items']):
        # pprint(item)
        iter_col = iter(range(5, 12))
        # print(next(iter_col))
        # if DISC is empty in that case assign other value
        if 'disc' not in item:
            item['disc'] = 0

        sheet.cell(row=sheet_row + i, column=1).value = sr_no
        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = item['part']
        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = item['desc']


        if item['qty'].__contains__('.'):
            qty = float(item['qty'])
        else:
            qty = int(item['qty'])

        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = qty

        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = float(item['unit'].replace(',',''))

        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = item['disc']
        sheet.cell(row=sheet_row + i, column=next(iter_col)).value = float(item['total'].replace(',',''))
        sr_no = sr_no + 1

    sheet.cell(row=sheet_row, column=1).value = 1
    sheet.cell(row=sheet_row, column=2).value = result['invoice_date']
    sheet.cell(row=sheet_row, column=3).value = result['invoice_no']
    sheet.cell(row=sheet_row, column=4).value = supplier_name
    # sheet.cell(row=2, column=9).value = person_data['po_number']
    # sheet.cell(row=2, column=10).value = person_data['address']
    # sheet.cell(row=2, column=11).value = person_data['terms']

    # temp_file = os.path.basename(filename)
    # out_filename = temp_file.rsplit('.', 1)[0] + '.xlsx'
    os.remove(prev_filename)
    print(os.path.join(dirname, f'table_{time_stamp}'))
    wb.save(f'{dirname}/table_{time_stamp}.xlsx')
    wb.close()
    sleep(0.5)


    print('Writing to Excel is done !!!')
